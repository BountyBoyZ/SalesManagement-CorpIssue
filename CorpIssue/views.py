from django.shortcuts import get_object_or_404, render, redirect
from django.http import JsonResponse, Http404
from django.urls import reverse
from django.conf import settings
from urllib.parse import urlencode
from CorpIssue.models import Corp, Invoice, Version, ConstValue, InvoiceTask, Task, Project, Team, RejectionDetails,InvoiceExcel
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, F, OuterRef, Sum, Case, When, FloatField
from django.db.models.functions import Coalesce
from django.db.models import F, Sum, OuterRef, Subquery
from django.db import connections
from io import BytesIO
import pandas as pd
from django.http import HttpResponse
from Utility.APIManager.Portal.register_document import v1 as register_document
from Utility.APIManager.Portal.send_document import ver1 as send_document
from Utility.Authentication.Utils import (
    V1_PermissionControl as permission_control,
    V1_get_data_from_token as get_token_data,
    V1_find_token_from_request as find_token
)
from django.core.paginator import Paginator
from django.db.models.functions import Cast
from django.db.models import IntegerField
from datetime import datetime
import logging
import json
import os
import openpyxl


logger = logging.getLogger(__name__)

@csrf_exempt
@permission_control
def list_corps(request, corp_code=None):
    # Check if user is sales manager
    token = find_token(request)
    user_nationalcode = get_token_data(token, "user_NationalCode")
    
    sales_manager_role = ConstValue.objects.filter(
        parent_code__code='StaticRoles',
        code='StaticRoles_SalesManager',
        value=user_nationalcode
    ).first()

    if not sales_manager_role:
        return render(request, "CorpIssue/error.html", {
            "error": "شما دسترسی لازم برای مشاهده این صفحه را ندارید"
        })

    corps = Corp.objects.all()
    corp = None
    latest_invoice = None
    status_value = None
    can_create_new_version = False
    create_new_version_url = None
    available_versions = []

    if corp_code:
        try:
            corp = Corp.objects.get(corp_code=corp_code)
        except Corp.DoesNotExist:
            return render(request, "CorpIssue/error.html", {"error": "شرکت مورد نظر یافت نشد."})
        latest_invoice = Invoice.objects.filter(corp_code=corp).order_by('-version_number').first()

        if not latest_invoice:
            status_value = 'نسخه ای موجود نیست'
            can_create_new_version = True
            available_versions = Version.objects.all().order_by('version_number')
        else:
            status_code = latest_invoice.status.code
            if status_code == "InvoiceStatus_Open":
                status_value = 'نسخه باز است'
                can_create_new_version = False
            elif status_code == "InvoiceStatus_Close":
                status_value = 'نسخه بسته است'
                can_create_new_version = True
                create_new_version_url = reverse('select_version', args=[corp_code])
            elif status_code == "InvoiceStatus_Review":
                status_value = 'در حال بررسی'
                can_create_new_version = False
            elif status_code == "InvoiceStatus_SentToSalesManager":
                status_value = 'ارسال شده به مدیر فروش'
                can_create_new_version = False
            elif status_code == "InvoiceStatus_SentToCustomer":
                status_value = 'ارسال شده به مشتری'
                can_create_new_version = False
            else:
                status_value = status_code
                can_create_new_version = False

    success_message = request.GET.get('success_message', '')

    return render(request, "CorpIssue/corp_list.html", {
        'corps': corps,
        'corp': corp,
        'latest_invoice': latest_invoice,
        'status_value': status_value,
        'can_create_new_version': can_create_new_version,
        'create_new_version_url': create_new_version_url,
        'selected_corp_code': corp_code,
        'available_versions': available_versions,
        'success_message': success_message,
        'help_doc': 'CorpIssue/Help Docs/Help 1.pdf'
    })


@csrf_exempt
@permission_control
def select_version(request, corp_code):
    """Select a new version for the corp and execute related actions."""
    token = find_token(request)
    user_nationalcode = get_token_data(token, "user_NationalCode")

    # Check if user is sales manager
    sales_manager_role = ConstValue.objects.filter(
        parent_code__code='StaticRoles',
        code='StaticRoles_SalesManager',
        value=user_nationalcode
    ).first()

    if not sales_manager_role:
        return render(request, "CorpIssue/error.html", {
            "error": "شما دسترسی لازم برای مشاهده این صفحه را ندارید"
        })

    try:
        corp = Corp.objects.get(corp_code=corp_code)
    except Corp.DoesNotExist:
        return render(request, "CorpIssue/error.html", {"error": "شرکت مورد نظر یافت نشد."})

    if request.method == 'POST':
        selected_version_number = request.POST.get('selected_version')
        selected_version = get_object_or_404(Version, version_number=selected_version_number)

        try:
            # Step 1: Insert a new record into the Invoice table
            status_review = ConstValue.objects.get(code="InvoiceStatus_Review")  # Get "Review" status

            invoice = Invoice.objects.create(
                corp_code=corp,
                version_number=selected_version,
                status=status_review,
                
                created_at=datetime.now(),
                updated_at=datetime.now(),
                send_at=datetime.now()
            )
            # Step 2: Create a new document for the selected version
            register_document_response = register_document(
                app_doc_id=invoice.id,
                priority='عادی',
                doc_state='بررسی توسط معاون محصول ',
                document_title=f"{selected_version_number} صورت حساب نسخه  {corp_code} شرکت ",
                app_code='CorpIssue',
                owner=user_nationalcode 
            )

            doc_id = register_document_response["data"]["id"]
            invoice.doc_id=doc_id

            # Step 3: Run the stored procedure for data transfer 
            try:
                with connections['default'].cursor() as cursor:
                    cursor.execute("EXEC CorpIssue_TransferInvoiceRecords @CorpCode = %s, @VersionNumber = %s,@Creator = %s,@Doc_id =%s ", 
                                [corp_code, selected_version_number,user_nationalcode,doc_id])
                    if cursor.rowcount == -1:  # Error occurred
                        invoice.delete()
                        return render(request, "CorpIssue/corp_list.html", {
                            'corp': corp,
                            'available_versions': Version.objects.all().order_by('version_number'),
                            'error': "خطا در انتقال اطلاعات"
                        })
            except Exception as e:
                invoice.delete()
                logger.error(f"Error in stored procedure execution: {str(e)}")
                return render(request, "CorpIssue/error.html", {
                    "error": "خطا در اجرای عملیات: " + str(e)
                })

            # Step 4: Retrieve "معاون محصول" ID
            product_assistant_id = ConstValue.objects.get(code="StaticRoles_ProductAssistant").value        

            # Step 5: Send the document to "معاون محصول"
            send_document_response = send_document(
                doc_id=register_document_response['data']['id'],
                sender=user_nationalcode,
                inbox_owners=[product_assistant_id]
            )

            # Step 6: Update invoice status
            invoice.status = ConstValue.objects.get(code='InvoiceStatus_Review')
            invoice.save()
            # Add success message to the context
            return redirect(reverse('list_corps', kwargs={'corp_code': corp_code}) + '?success_message=درخواست به معاون محصول ارسال گردید')

        except Exception as e:
            logger.error(f"Error in select_version: {e}")
            return render(request, "CorpIssue/error.html", {"error": str(e)})

        return redirect('list_corps', corp_code=corp_code)

    return render(request, "CorpIssue/corp_list.html", {
        'corp': corp,
        'available_versions': Version.objects.all().order_by('version_number')
    })


def get_query_string(request, page_number):
    """Generate query string preserving all parameters except page."""
    query_dict = request.GET.copy()
    query_dict['page'] = page_number
    return query_dict.urlencode(safe='/')


@csrf_exempt
@permission_control
def invoice_tasks(request, invoice_id):
    """Show all invoice tasks related to the invoice ID."""
    token = find_token(request)
    user_nationalcode = get_token_data(token, "user_NationalCode")

    # Check if user belongs to a valid team (POD or PMA)
    user_team = Team.objects.filter(manager=user_nationalcode).first()
    if not user_team:
        return render(request, "CorpIssue/error.html", {"error": "شما دسترسی لازم برای مشاهده این صفحه را ندارید"})

    try:
        invoice = Invoice.objects.get(id=invoice_id)
    except Invoice.DoesNotExist:
        return render(request, "CorpIssue/error.html", {"error": "صورت حساب مورد نظر یافت نشد."})

    corp = invoice.corp_code
    version_number = invoice.version_number.version_number

    # Assign default status to tasks without status
    pending_status = ConstValue.objects.get(code='InvoiceTaskStatus_PendingByProductAssistant')
    for task in InvoiceTask.objects.filter(invoice=invoice, status__isnull=True):
        task.status = pending_status
        task.save()

    # Fetch invoice tasks based on user role
    if user_team.team_code == 'POD':
        is_product_assistant = True
        is_team_manager = False
        # Only show table if invoice is under review
        show_table = invoice.status.code == 'InvoiceStatus_Review'
        invoice_tasks = InvoiceTask.objects.filter(invoice=invoice) if show_table else InvoiceTask.objects.none()
    else:
        # Update team manager's view to include customer rejected tasks when invoice status is ReturnedToProjectManager
        filter_conditions = [
            'InvoiceTaskStatus_RejectedByProductAssistant',
            'InvoiceTaskStatus_Rejected'
        ]
        
        # Add RejectedByCustomer to filter conditions when invoice is in ReturnedToProjectManager status
        if invoice.status.code == 'InvoiceStatus_ReturnedToProjectManager':
            filter_conditions.append('InvoiceTaskStatus_RejectedByCustomer')
            filter_conditions.append('InvoiceTaskStatus_RejectedBySalesManager')

        invoice_tasks = InvoiceTask.objects.filter(
            invoice=invoice,
            task__project__team_code=user_team,
            status__code__in=filter_conditions
        )
        is_product_assistant = False
        is_team_manager = True
        show_table = True

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        invoice_tasks = invoice_tasks.filter(
            Q(task__task_id__icontains=search_query) |
            Q(task__task_title__icontains=search_query) |
            Q(task__task_kind__value__icontains=search_query) |
            Q(task__project__project_name__icontains=search_query) |
            Q(task__project__team_code__team_name__icontains=search_query) |
            Q(task__real_work_hours__icontains=search_query) |
            Q(invoice_work_hours__icontains=search_query) |
            Q(status__value__icontains=search_query)

        )

    # Filter functionality
    task_id = request.GET.get('task_id', '')
    task_title = request.GET.get('task_title', '')
    status_codes = request.GET.getlist('status')

    if task_id:
        invoice_tasks = invoice_tasks.filter(task__task_id__icontains=task_id)
    if task_title:
        invoice_tasks = invoice_tasks.filter(task__task_title__icontains=task_title)
    if status_codes:
        invoice_tasks = invoice_tasks.filter(status__code__in=status_codes)

    #task kind filter
    task_kind = request.GET.get('task_kind')
    if task_kind:
        invoice_tasks = invoice_tasks.filter(task__task_kind_id=task_kind)

    #real work hours filter
    real_work_hours_type = request.GET.get('real_work_hours_type')
    real_work_hours_min = request.GET.get('real_work_hours_min')
    real_work_hours_max = request.GET.get('real_work_hours_max')

    if real_work_hours_type and real_work_hours_min:
        if real_work_hours_type == 'gt':
            invoice_tasks = invoice_tasks.filter(task__real_work_hours__gte=real_work_hours_min)  # Changed from gt to gte
        elif real_work_hours_type == 'lt':
            invoice_tasks = invoice_tasks.filter(task__real_work_hours__lte=real_work_hours_min)  # Changed from lt to lte
        elif real_work_hours_type == 'eq':
            invoice_tasks = invoice_tasks.filter(task__real_work_hours=real_work_hours_min)
        elif real_work_hours_type == 'between' and real_work_hours_max:
            invoice_tasks = invoice_tasks.filter(
                task__real_work_hours__gte=real_work_hours_min,  # Changed from gt to gte
                task__real_work_hours__lte=real_work_hours_max   # Changed from lt to lte
            )

    #invoice work hours filter
    invoice_work_hours_type = request.GET.get('invoice_work_hours_type')
    invoice_work_hours_min = request.GET.get('invoice_work_hours_min')
    invoice_work_hours_max = request.GET.get('invoice_work_hours_max')

    if invoice_work_hours_type and invoice_work_hours_min:
        if invoice_work_hours_type == 'gt':
            invoice_tasks = invoice_tasks.filter(invoice_work_hours__gte=invoice_work_hours_min)  # Changed from gt to gte
        elif invoice_work_hours_type == 'lt':
            invoice_tasks = invoice_tasks.filter(invoice_work_hours__lte=invoice_work_hours_min)  # Changed from lt to lte
        elif invoice_work_hours_type == 'eq':
            invoice_tasks = invoice_tasks.filter(invoice_work_hours=invoice_work_hours_min)
        elif invoice_work_hours_type == 'between' and invoice_work_hours_max:
            invoice_tasks = invoice_tasks.filter(
                invoice_work_hours__gte=invoice_work_hours_min,  # Changed from gt to gte
                invoice_work_hours__lte=invoice_work_hours_max   # Changed from lt to lte
            )

    # Work hours ratio filter
    work_hours_ratio_type = request.GET.get('work_hours_ratio_type')

    if work_hours_ratio_type and work_hours_ratio_type != 'all':
        invoice_tasks = invoice_tasks.annotate(
            work_hours_ratio=Case(
                When(task__real_work_hours=0, then=None),
                default=(F('invoice_work_hours') * 100.0 / F('task__real_work_hours')) - 100.0,  # Subtract 100
                output_field=FloatField(),
            )
        )

        ratio_value = float(request.GET.get('work_hours_ratio_value', 0))
        ratio_max = float(request.GET.get('work_hours_ratio_max', 0))

        if work_hours_ratio_type == 'gt':
            invoice_tasks = invoice_tasks.filter(
                work_hours_ratio__isnull=False,
                work_hours_ratio__gte=ratio_value
            )
        elif work_hours_ratio_type == 'lt':
            invoice_tasks = invoice_tasks.filter(
                work_hours_ratio__isnull=False,
                work_hours_ratio__lte=ratio_value
            )
        elif work_hours_ratio_type == 'eq':
            invoice_tasks = invoice_tasks.filter(
                work_hours_ratio__isnull=False,
                work_hours_ratio=ratio_value
            )
        elif work_hours_ratio_type == 'between' and ratio_max:
            invoice_tasks = invoice_tasks.filter(
                work_hours_ratio__isnull=False,
                work_hours_ratio__gte=ratio_value,
                work_hours_ratio__lte=ratio_max
            )

    # Implementation ratio filter
    implementation_ratio_type = request.GET.get('implementation_ratio_type')
    implementation_ratio_value = request.GET.get('implementation_ratio_value')

    if implementation_ratio_type and implementation_ratio_type != 'all' and implementation_ratio_value:
        # Subquery for implementation hours (Analysis + Design + Cognition + Implementation)
        implementation_hours = Task.objects.filter(
            parent_task=OuterRef('task_id'),
            task_kind__code__in=[
                'TaskKind_SubAnalysis',
                'TaskKind_SubDesign',
                'TaskKind_SubCognition',
                'TaskKind_SubImplementation'
            ]
        ).values('parent_task').annotate(
            total_hours=Coalesce(Sum('real_work_hours'), 0)
        ).values('total_hours')

        # Subquery for debug/test hours (Test + Debug + Problem)
        debug_test_hours = Task.objects.filter(
            parent_task=OuterRef('task_id'),
            task_kind__code__in=[
                'TaskKind_SubTest',
                'TaskKind_SubDebug',
                'TaskKind_SubProblem'
            ]
        ).values('parent_task').annotate(
            total_hours=Coalesce(Sum('real_work_hours'), 0)
        ).values('total_hours')

        # Fix the ratio calculation to be (implementation / debug_test * 100) - 100
        invoice_tasks = invoice_tasks.annotate(
            implementation_total=Coalesce(Subquery(implementation_hours), 0),
            debug_test_total=Coalesce(Subquery(debug_test_hours), 0),
            ratio=Case(
                When(debug_test_total=0, then=None),
                default=(F('implementation_total') * 100.0 / F('debug_test_total')) - 100.0,  # Subtract 100
                output_field=FloatField(),
            )
        )

        # Convert input percentage to float
        ratio_value = float(implementation_ratio_value)

        # Apply filters with correct comparison
        if implementation_ratio_type == 'gt':
            invoice_tasks = invoice_tasks.filter(
                ratio__isnull=False,
                ratio__gte=ratio_value
            )
        elif implementation_ratio_type == 'lt':
            invoice_tasks = invoice_tasks.filter(
                ratio__isnull=False,
                ratio__lte=ratio_value
            )
        elif implementation_ratio_type == 'eq':
            invoice_tasks = invoice_tasks.filter(
                ratio__isnull=False,
                ratio=ratio_value
            )

    # Team filter
    selected_teams = request.GET.getlist('team')
    if selected_teams:
        invoice_tasks = invoice_tasks.filter(
            task__project__team_code__team_code__in=selected_teams
        )

    # Project filter
    selected_projects = request.GET.getlist('project')
    if selected_projects:
        invoice_tasks = invoice_tasks.filter(
            task__project_id__in=selected_projects
        )

    # Get task statuses for the filter
    task_statuses = ConstValue.objects.filter(parent_code_id=22)
    selected_statuses = status_codes

    # Define status map
    status_map = {
        'InvoiceTaskStatus_Approved': {'label': 'تایید شده', 'class': 'status-approved'},
        'InvoiceTaskStatus_ApprovedByProductAssistant': {'label': 'تایید شده توسط معاون محصول', 'class': 'status-approved'},
        'InvoiceTaskStatus_ApprovedByProjectManager': {'label': 'تایید شده توسط مدیر پروژه', 'class': 'status-approved-pm'},
        'InvoiceTaskStatus_Deleted': {'label': 'حذف شده', 'class': 'status-rejected'},
        'InvoiceTaskStatus_Rejected': {'label': 'رد شده', 'class': 'status-rejected'},
        'InvoiceTaskStatus_RejectedByProductAssistant': {'label': 'رد شده توسط معاون محصول', 'class': 'status-rejected'},
        'InvoiceTaskStatus_PendingByProductAssistant': {'label': 'در انتظار معاون محصول', 'class': 'status-pending'},
        'InvoiceTaskStatus_ReturnedByProjectManager': {'label': 'برگشت داده شده', 'class': 'status-returned'}
    }

    tasks_data = []

    # Fetch related data in bulk to optimize performance
    invoice_tasks = invoice_tasks.select_related('task__project', 'status')
        # Pagination
    
    paginator = Paginator(invoice_tasks, 20)  # 20 tasks per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    for invoice_task in page_obj:
        task = invoice_task.task
        project = task.project
        team = project.team_code  # Assuming team_code is a string in project

        # Determine status class for UI styling
        status_code = invoice_task.status.code
        status_class = {
            'InvoiceTaskStatus_Approved': 'status-approved',
            'InvoiceTaskStatus_ApprovedByProductAssistant': 'status-approved',
            'InvoiceTaskStatus_ApprovedByProjectManager': 'status-approved-pm',
            'InvoiceTaskStatus_Deleted': 'status-rejected',
            'InvoiceTaskStatus_Rejected': 'status-rejected',
            'InvoiceTaskStatus_RejectedByProductAssistant': 'status-rejected',
            'InvoiceTaskStatus_PendingByProductAssistant': 'status-pending',
            'InvoiceTaskStatus_ReturnedByProjectManager': 'status-returned'
        }.get(status_code, 'status-other')

        

        # Fetch RejectionDetails based on invoice status
        if invoice.status.code == 'InvoiceStatus_ReturnedToProjectManager':
            # Fetch the first (oldest) rejection detail
            rejection_details = RejectionDetails.objects.filter(invoice_task=invoice_task).first()
        else:
            # Fetch the latest rejection detail
            rejection_details = RejectionDetails.objects.filter(invoice_task=invoice_task).order_by('-created_at').first()

        rejection_reason = rejection_details.explanation if rejection_details else ''
        rejection_title = rejection_details.const_value.value if rejection_details else ''

        tasks_data.append({
            'task_id': task.task_id,
            'task_title': task.task_title,
            'task_type': task.task_kind.value,
            'project': project.project_name,
            'team': team.team_code,  # Assuming team_code is the correct field
            'real_work_hours': task.real_work_hours_display,
            'invoice_work_hours': invoice_task.invoice_work_hours_display,
            'status': invoice_task.status.value,
            'status_class': status_class,
            'rejection_reason': rejection_reason,
            'rejection_title': rejection_title,
            'task_statuses': task_statuses,
            'selected_statuses': selected_statuses,
            'help_doc': 'CorpIssue/Help Docs/Help 1.pdf'
        })
    # Get task types, teams, and statuses for the dropdowns
    task_types = ConstValue.objects.filter(parent_code__code='TaskKind')
    # Modify the teams query to only get teams with projects in this invoice
    teams = Team.objects.filter(
        project__task__invoicetask__invoice=invoice
    ).distinct().select_related()  # Add select_related for performance


    statuses = ConstValue.objects.filter(parent_code__code='InvoiceTaskStatus')


    # Get rejection reasons for the dropdown
    rejection_reasons = ConstValue.objects.filter(parent_code__code='InvoiceTaskStatus_RejectedByProductAssistant')

    # Get selected teams and their projects
    selected_teams = request.GET.getlist('team')
    selected_projects = request.GET.getlist('project')
    
    # Get projects for selected teams
    team_projects = []
    if selected_teams:
        team_projects = Project.objects.filter(
            team_code__team_code__in=selected_teams
        ).values('id', 'project_name', 'team_code__team_code')
        # Rename for template compatibility
        team_projects = [
            {
                'id': p['id'],
                'project_name': p['project_name'],
                'team_code': p['team_code__team_code'],
            }
            for p in team_projects
        ]

    success_message = request.GET.get('success_message', '')

    all_projects = Project.objects.select_related('team_code').all()

    return render(request, "CorpIssue/InvoiceTask_form.html", {
        'tasks_data': tasks_data,
        'is_product_assistant': is_product_assistant,
        'is_team_manager': is_team_manager,
        'rejection_reasons': rejection_reasons,
        'task_types': task_types,
        'teams': teams,
        'statuses': statuses,
        'corp_name': corp.corp_name,
        'corp_code': corp.corp_code,
        'version_number': version_number,
        'invoice': invoice,
        'page_obj': page_obj,
        'status_map': status_map,
        'selected_statuses': status_codes,
        'selected_teams': selected_teams,
        'team_projects': list(team_projects),
        'selected_projects': selected_projects,
        'query_string': get_query_string(request, page_obj.number),
        'help_doc': 'CorpIssue/Help Docs/Help 2.pdf',
        'success_message': success_message,
        'show_table': show_table,
        'all_projects': list(all_projects)
    })


def send_to_sales_manager(invoice, user_nationalcode):
    """Helper function to send document to sales manager."""
    try:
        # Get sales manager's username from ConstValue
        sales_manager = ConstValue.objects.get(code='StaticRoles_SalesManager').value
        
        # Send document using ver1
        send_document_response = send_document(
            doc_id=invoice.doc_id,
            sender=user_nationalcode,
            inbox_owners=[sales_manager]
        )

        # Check if document was sent successfully
        if not all(response.get('msg') == 'success' for response in send_document_response.values()):
            return False, 'Failed to send document to sales manager'
        return True, None
    except Exception as e:
        logger.error(f"Error sending document to sales manager: {e}")
        return False, str(e)

@csrf_exempt
@permission_control
def approve_task(request, task_id):
    token = find_token(request)
    user_nationalcode = get_token_data(token, "user_NationalCode")

    try:
        invoice_task = InvoiceTask.objects.get(task_id=task_id)
    except InvoiceTask.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'تسک مورد نظر یافت نشد.'})

    user_team = Team.objects.filter(manager=user_nationalcode).first()
    if not user_team:
        return JsonResponse({'success': False, 'error': 'شما دسترسی لازم برای انجام این عملیات را ندارید.'})

    data = json.loads(request.body) if request.body else {}
    response_text = data.get('response', '')

    if user_team.team_code == 'POD':
        # Keep existing POD logic unchanged
        approved_status = ConstValue.objects.get(code='InvoiceTaskStatus_ApprovedByProductAssistant')
        invoice_task.status = approved_status
        invoice_task.save()
        
        team_manager = invoice_task.task.project.team_code.manager
        if team_manager:
            send_document(
                doc_id=invoice_task.invoice.doc_id,
                sender=request.user.username,
                inbox_owners=[team_manager]
            )
        
        # Check if all tasks are approved
        invoice = invoice_task.invoice
        if not InvoiceTask.objects.filter(invoice=invoice).exclude(status=approved_status).exists():
            sent_to_sales_manager_status = ConstValue.objects.get(code='InvoiceStatus_SentToSalesManager')
            invoice.status = sent_to_sales_manager_status
            invoice.save()

            success, error = send_to_sales_manager(invoice, user_nationalcode)
            if not success:
                return JsonResponse({'success': False, 'error': error})

    else:
        # Handle both existing team manager logic and customer rejection case
        if invoice_task.status.code == 'InvoiceTaskStatus_RejectedByCustomer':
            if not response_text:
                return JsonResponse({
                    'success': False,
                    'error': 'پاسخ فناوران الزامی است'
                })

            # Find all rejection details for this task, ordered by created_at
            rejections = list(RejectionDetails.objects.filter(invoice_task=invoice_task).order_by('created_at'))

            # Find the latest insurance comment (RejectedByCustomer) that does NOT have a later ApprovedByProjectManager after it
            latest_unanswered = None
            for i, rej in enumerate(rejections):
                if rej.const_value.code == 'InvoiceTaskStatus_RejectedByCustomer':
                    # Check if there is an ApprovedByProjectManager after this
                    has_response = False
                    for later in rejections[i+1:]:
                        if later.const_value.code == 'InvoiceTaskStatus_ApprovedByProjectManager':
                            has_response = True
                            break
                    if not has_response:
                        latest_unanswered = rej

            # Now, when the sales manager submits a response, attach it to this latest_unanswered
            if latest_unanswered:
                latest_unanswered.explanation = response_text
                latest_unanswered.save()

            # Create rejection details with response
            approved_status = ConstValue.objects.get(code='InvoiceTaskStatus_ApprovedByProjectManager')
            RejectionDetails.objects.create(
                const_value=approved_status,
                invoice_task=invoice_task,
                explanation=response_text,
                rejected_by=user_nationalcode,
                created_by=user_nationalcode
            )

            # Update task status
            invoice_task.status = approved_status
            invoice_task.save()

            # Only auto-transition if invoice is NOT in ReturnedToProjectManager
            if invoice_task.invoice.status.code != 'InvoiceStatus_ReturnedToProjectManager':
                remaining_rejected = InvoiceTask.objects.filter(
                    invoice=invoice_task.invoice,
                    status__code='InvoiceTaskStatus_RejectedByCustomer'
                ).exists()

                if not remaining_rejected:
                    # Change invoice status to SentToSalesManager
                    sent_to_sales_status = ConstValue.objects.get(code='InvoiceStatus_SentToSalesManager')
                    invoice_task.invoice.status = sent_to_sales_status
                    invoice_task.invoice.save()

                    # Send document to sales manager
                    sales_manager = ConstValue.objects.get(code='StaticRoles_SalesManager')
                    if sales_manager and sales_manager.value:
                        send_document(
                            doc_id=invoice_task.invoice.doc_id,
                            sender=user_nationalcode,
                            inbox_owners=[sales_manager.value]
                        )
        else:
            approved_status = ConstValue.objects.get(code='InvoiceTaskStatus_ApprovedByProjectManager')
            invoice_task.status = approved_status
            invoice_task.save()

            # Always create a new RejectionDetails for نظر فناوران
            response_text = data.get('response', '')
            if response_text:
                RejectionDetails.objects.create(
                    const_value=approved_status,
                    invoice_task=invoice_task,
                    explanation=response_text,
                    rejected_by=user_nationalcode,
                    created_by=user_nationalcode
                )

            sales_manager = ConstValue.objects.get(code='StaticRoles_SalesManager')
            if sales_manager and sales_manager.value:
                send_document(
                    doc_id=invoice_task.invoice.doc_id,
                    sender=user_nationalcode,
                    inbox_owners=[sales_manager.value]
                )

    return JsonResponse({'success': True})

@csrf_exempt
@permission_control
def reject_task(request, task_id):
    """Reject the task and send document to team manager."""
    token = find_token(request)
    user_nationalcode = get_token_data(token, "user_NationalCode")

    try:
        data = json.loads(request.body)
        reason_id = data.get('reason_id')
        explanation = data.get('explanation')

        try:
            invoice_task = InvoiceTask.objects.get(task_id=task_id)
        except InvoiceTask.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'تسک مورد نظر یافت نشد.'})

        user_team = Team.objects.filter(manager=user_nationalcode).first()

        if user_team.team_code == 'POD':
            rejected_status = ConstValue.objects.get(code='InvoiceTaskStatus_RejectedByProductAssistant')
            
            # Get the document ID from the invoice
            doc_id = invoice_task.invoice.doc_id
            
            # Get team manager username
            team_manager = invoice_task.task.project.team_code.manager
            
            # Send document to team manager using ver1
            send_document_response = send_document(
                doc_id=doc_id,
                sender=user_nationalcode,
                inbox_owners=[team_manager]
            )

            # Check if document was sent successfully
            if not all(response.get('msg') == 'success' for response in send_document_response.values()):
                return JsonResponse({
                    'success': False, 
                    'error': 'Failed to send document to team manager'
                })

            # Always create a new rejection detail record
            rejection_reason = get_object_or_404(ConstValue, id=reason_id)
            RejectionDetails.objects.create(
                const_value=rejection_reason,
                invoice_task=invoice_task,
                explanation=explanation,
                rejected_by=user_nationalcode,
                created_by=user_nationalcode
            )

        elif user_team.team_code == 'PMA':
            rejected_status = ConstValue.objects.get(code='InvoiceTaskStatus_RejectedByProjectManager')
        else:
            return JsonResponse({'success': False, 'error': 'Unauthorized'})

        # Update task status
        invoice_task.status = rejected_status
        invoice_task.save()

        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Error in reject_task: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
@permission_control
def rejection_details(request, task_id):
    """Fetch latest rejection details for the given task ID."""
    try:
        token = find_token(request)
        user_nationalcode = get_token_data(token, "user_NationalCode")

        # Allow both team managers and sales manager
        user_team = Team.objects.filter(manager=user_nationalcode).first()
        is_sales_manager = ConstValue.objects.filter(
            parent_code__code='StaticRoles',
            code='StaticRoles_SalesManager',
            value=user_nationalcode
        ).exists()

        if not user_team and not is_sales_manager:
            return JsonResponse({'success': False, 'error': 'شما دسترسی لازم برای مشاهده این صفحه را ندارید.'})

        # Always show the latest rejection detail
        rejection_detail = RejectionDetails.objects.filter(
            invoice_task__task_id=task_id
        ).order_by('-created_at').first()

        if rejection_detail:
            data = {
                'success': True,
                'rejection_title': rejection_detail.const_value.value,
                'rejection_explanation': rejection_detail.explanation
            }
        else:
            data = {'success': False, 'error': 'No rejection details found'}

    except Exception as e:
        logger.error(f"Error in rejection_details: {e}")
        data = {'success': False, 'error': str(e)}

    return JsonResponse(data)

@csrf_exempt
@permission_control
def get_team_projects(request):
    """Get projects for selected teams."""
    try:
        team_codes = request.GET.get('teams', '').split(',')
        if not team_codes or team_codes[0] == '':
            return JsonResponse({
                'success': True,
                'projects': []
            })

        # Update query to get all necessary fields
        projects = Project.objects.filter(
            team_code__team_code__in=team_codes
        ).select_related('team_code').values(
            'id',
            'project_name',
            'team_code__team_code',  # Get team_code from joined Team model
            'team_code__team_name'   # Get team_name for reference
        ).distinct()

        # Transform the data 
        projects_data = []
        for project in projects:
            if project['team_code__team_code']:
                projects_data.append({
                    'id': project['id'],
                    'project_name': project['project_name'],
                    'team_code': project['team_code__team_code'],
                    'team_name': project['team_code__team_name'],
                })

        print("Selected teams:", team_codes)        # Debug log
        print("Found projects:", projects_data)     # Debug log

        return JsonResponse({
            'success': True,
            'projects': projects_data
        })
    except Exception as e:
        print(f"Error in get_team_projects: {str(e)}")  # Debug log
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
@csrf_exempt
@permission_control
def sub_tasks(request, task_id):
    """Fetch subtasks for the given task ID."""
    try:
        try:
            task = Task.objects.get(task_id=task_id)
        except Task.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'تسک مورد نظر یافت نشد.'})

        sub_tasks = task.sub_tasks.all()
        sub_tasks_data = []
        for sub_task in sub_tasks:
            sub_tasks_data.append({
                'task_kind': sub_task.task_kind.value,  # Use the value field for translation
                'real_work_hours': sub_task.real_work_hours_display 
            })
        data = {'success': True, 'sub_tasks': sub_tasks_data}
    except Exception as e:
        logger.error(f"Error in sub_tasks: {e}")
        data = {'success': False, 'error': str(e)}

    return JsonResponse(data)

@csrf_exempt
@permission_control
def approve_all_tasks(request, invoice_id):
    """Approve all tasks for the given invoice ID."""
    token = find_token(request)
    user_nationalcode = get_token_data(token, "user_NationalCode")

    try:
        try:
            invoice = Invoice.objects.get(id=invoice_id)
        except Invoice.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'صورت حساب مورد نظر یافت نشد.'})

        approved_status = ConstValue.objects.get(code='InvoiceTaskStatus_ApprovedByProductAssistant')
        sent_to_sales_manager_status = ConstValue.objects.get(code='InvoiceStatus_SentToSalesManager')

        # Approve all tasks
        InvoiceTask.objects.filter(invoice=invoice).update(status=approved_status)

        # Update invoice status
        invoice.status = sent_to_sales_manager_status
        invoice.save()

        # Send document to sales manager
        success, error = send_to_sales_manager(invoice, user_nationalcode)
        if not success:
            return JsonResponse({'success': False, 'error': error})

        # Redirect with success message
        return JsonResponse({
            'success': True,
            'redirect_url': reverse('invoice_tasks', args=[invoice_id]) + '?success_message=فرم نسخه برای مدیر فروش ارسال گردید.'
        })
    except Exception as e:
        logger.error(f"Error in approve_all_tasks: {e}")
        return JsonResponse({'success': False, 'error': str(e)})
    

@csrf_exempt
@permission_control
def sales_manager_view(request, invoice_id):
    """View for the sales manager to handle invoice tasks."""
    token = find_token(request)
    user_nationalcode = get_token_data(token, "user_NationalCode")
    
    # Check if the user is a sales manager
    sales_manager_role = ConstValue.objects.filter(
        parent_code__code='StaticRoles',
        code='StaticRoles_SalesManager',
        value=user_nationalcode
    ).first()

    if not sales_manager_role:
        return render(request, "CorpIssue/sales_manager_form.html", {
            "error": "شما دسترسی لازم برای مشاهده این صفحه را ندارید"
        })

    try:
        invoice = Invoice.objects.get(id=invoice_id)
    except Invoice.DoesNotExist:
        return render(request, "CorpIssue/error.html", {"error": "صورت حساب مورد نظر یافت نشد."})

    corp = invoice.corp_code
    version_number = invoice.version_number.version_number

    # Handle download request
    if request.method == 'POST' and 'download' in request.POST:
        if invoice.status.code in ['InvoiceStatus_SentToSalesManager', 'InvoiceStatus_SentToCustomer']:
            invoice_tasks = InvoiceTask.objects.filter(invoice=invoice).select_related(
                'task__project', 
                'task__task_kind', 
                'status'
            )

            tasks_data = []
            max_n = 1  # Track the maximum N across all tasks

            # First pass: collect all comments/responses and determine max N
            task_comments = {}
            for idx, invoice_task in enumerate(invoice_tasks, 1):
                task = invoice_task.task
                project = task.project
                team = project.team_code

                # Get rejection details if any
                rejection_details = list(RejectionDetails.objects.filter(
                    invoice_task=invoice_task
                ).order_by('created_at'))

                insurance_comments = []
                dev_responses = []

                # Track the index of each insurance comment
                insurance_indices = []
                for i, rejection in enumerate(rejection_details):
                    if rejection.const_value.code == 'InvoiceTaskStatus_RejectedByCustomer':
                        insurance_comments.append(rejection.explanation)
                        insurance_indices.append(i)

                # For each insurance comment, find the latest dev response after it and before the next insurance comment
                for n, start_idx in enumerate(insurance_indices):
                    end_idx = insurance_indices[n + 1] if n + 1 < len(insurance_indices) else len(rejection_details)
                    latest_response = ''
                    for j in range(start_idx + 1, end_idx):
                        r = rejection_details[j]
                        if r.const_value.code == 'InvoiceTaskStatus_ApprovedByProjectManager':
                            latest_response = r.explanation
                    dev_responses.append(latest_response)

                # Pad lists to the same length
                max_length = max(len(insurance_comments), len(dev_responses))
                insurance_comments += [''] * (max_length - len(insurance_comments))
                dev_responses += [''] * (max_length - len(dev_responses))

                # Save for second pass
                task_comments[task.task_id] = {
                    'idx': idx,
                    'task': task,
                    'project': project,
                    'team': team,
                    'invoice_task': invoice_task,
                    'insurance_comments': insurance_comments,
                    'dev_responses': dev_responses,
                    'max_length': max_length
                }
                if max_length > max_n:
                    max_n = max_length

            # Second pass: build the DataFrame rows with N+1 columns
            for task_id, data in task_comments.items():
                idx = data['idx']
                task = data['task']
                project = data['project']
                team = data['team']
                invoice_task = data['invoice_task']
                insurance_comments = data['insurance_comments']
                dev_responses = data['dev_responses']
                max_length = data['max_length']

                # Pad to max_n for all tasks
                insurance_comments = insurance_comments + [''] * (max_n - max_length + 1)  # +1 for N+1
                dev_responses = dev_responses + [''] * (max_n - max_length + 1)

                # Find the first rejection detail for this task with const_value code 'InvoiceTaskStatus_RejectedByCustomer'
                # OR whose parent_code is 'InvoiceTaskStatus_RejectedByCustomer'
                rejected_by_customer_const = ConstValue.objects.get(code='InvoiceTaskStatus_RejectedByCustomer')
                initial_rejection_detail = RejectionDetails.objects.filter(
                    invoice_task=invoice_task
                ).filter(
                    Q(const_value__code='InvoiceTaskStatus_RejectedByCustomer') |
                    Q(const_value__parent_code=rejected_by_customer_const)
                ).order_by('created_at').first()
                initial_rejection_title = initial_rejection_detail.const_value.value if initial_rejection_detail else ''

                task_data = {
                    'ردیف': idx,
                    'شناسه تسک': task.task_id,
                    'عنوان تسک': task.task_title,
                    'نوع تسک': task.task_kind.value,
                    'پروژه': project.project_name,
                    'تیم': team.team_name,
                    'کارکرد واقعی': task.real_work_hours_display,
                    'کارکرد اعلامی': invoice_task.invoice_work_hours_display,
                    'وضعیت': invoice_task.status.value,
                    'دلیل عدم تایید': initial_rejection_title, 
                    'نظر شرکت بیمه': insurance_comments[0] if insurance_comments else '',
                    'پاسخ فناوران': dev_responses[0] if dev_responses else ''
                }

                # Add additional comments and responses with numbering
                for i in range(1, max_n + 1):  # +1 for N+1
                    task_data[f'نظر شرکت بیمه {i+1}'] = insurance_comments[i]
                    task_data[f'پاسخ فناوران {i+1}'] = dev_responses[i]

                tasks_data.append(task_data)

            df = pd.DataFrame(tasks_data)
            excel_file = BytesIO()
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='صورت حساب')
                # Load the Release_Comment.xlsx and append as second sheet
                release_comment_path = os.path.join(
                    settings.BASE_DIR, "static", "CorpIssue", "Help Docs", "Release_Comment.xlsx"
                )
                if os.path.exists(release_comment_path):
                    release_wb = openpyxl.load_workbook(release_comment_path)
                    release_ws = release_wb.active
                    # Create a new worksheet in the current writer
                    ws = writer.book.create_sheet(title='راهنمای انتشار')
                    for row in release_ws.iter_rows(values_only=True):
                        ws.append(row)
            excel_file.seek(0)

            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=invoice_{corp.corp_code}_{version_number}.xlsx'
            return response

    # Handle file upload
    elif request.method == 'POST' and 'upload' in request.FILES:
        if invoice.status.code == 'InvoiceStatus_SentToCustomer':
            try:
                uploaded_file = request.FILES['upload']
                upload_dir = os.path.join(settings.BASE_DIR, 'static', 'CorpIssue', 'Excel_Upload')
                os.makedirs(upload_dir, exist_ok=True)
                file_name = f'invoice_{corp.corp_code}_{version_number}.xlsx'
                file_path = os.path.join(upload_dir, file_name)
                with open(file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)
                df = pd.read_excel(file_path)

                # Normalize column names (strip, remove ZWNJ, etc.)
                df.columns = [str(col).strip().replace('\u200c', '').replace('\u202a', '').replace('\u202c', '') for col in df.columns]
                # Find the exact column name for "دلیل عدم تایید"
                reason_col = next((col for col in df.columns if 'دلیل عدم تایید' in col.replace('\u200c', '').replace('\u202a', '').replace('\u202c', '')), None)
                if not reason_col:
                    return JsonResponse({'success': False, 'error': 'ستون دلیل عدم تایید در فایل اکسل یافت نشد.'})

                rejected_by_customer_const = ConstValue.objects.get(code='InvoiceTaskStatus_RejectedByCustomer')
                parent_id = rejected_by_customer_const.id

                # Fetch all possible rejection reasons for this parent, normalize for comparison
                possible_reasons = ConstValue.objects.filter(parent_code_id=parent_id)
                normalized_db_reasons = {
                    str(val.value).strip().replace('\u200c', '').replace('\u202a', '').replace('\u202c', ''): val
                    for val in possible_reasons
                }

                for _, row in df.iterrows():
                    task_id = row['شناسه تسک']
                    # Normalize the Excel value for comparison
                    rejection_reason = row.get(reason_col)
                    normalized_excel_reason = str(rejection_reason).strip().replace('\u200c', '').replace('\u202a', '').replace('\u202c', '') if rejection_reason and not pd.isna(rejection_reason) else None

                    insurance_comments = [row[col] for col in df.columns if col.startswith('نظر شرکت بیمه')]
                    latest_insurance_comment = ''
                    for comment in reversed(insurance_comments):
                        if isinstance(comment, str) and comment.strip():
                            latest_insurance_comment = comment
                            break

                    if normalized_excel_reason:
                        invoice_task = InvoiceTask.objects.get(task_id=task_id)
                        # Try to find a match in normalized db reasons
                        rejection_const = normalized_db_reasons.get(normalized_excel_reason, rejected_by_customer_const)
                        # Create RejectionDetails with the found or fallback const_value
                        RejectionDetails.objects.create(
                            const_value=rejection_const,
                            invoice_task=invoice_task,
                            explanation=latest_insurance_comment if latest_insurance_comment else '',
                            rejected_by=user_nationalcode,
                            created_by=user_nationalcode
                        )
                        # Update task status
                        invoice_task.status = rejected_by_customer_const
                        invoice_task.save()
                        InvoiceExcel.objects.create(invoice=invoice, file_path=file_path)
                        # Send document to team manager
                        team_manager = invoice_task.task.project.team_code.manager
                        send_document(
                            doc_id=invoice.doc_id,
                            sender=user_nationalcode,
                            inbox_owners=[team_manager]
                        )
                return JsonResponse({'success': True, 'message': 'فایل با موفقیت آپلود شد'})
            except Exception as e:
                logger.error(f"Error processing file: {str(e)}")
                return JsonResponse({'success': False, 'error': str(e)})

    if invoice.status.code == 'InvoiceStatus_ReturnedToProjectManager':
        allowed_statuses = [
            'InvoiceTaskStatus_RejectedByCustomer',
            'InvoiceTaskStatus_ApprovedByProjectManager',
            'InvoiceTaskStatus_RejectedBySalesManager',
            'InvoiceTaskStatus_ApprovedBySalesManager'
        ]
        # Add tasks whose status parent is InvoiceTaskStatus_RejectedByCustomer_NotDone
        rejected_notdone_parent = ConstValue.objects.get(code='InvoiceTaskStatus_RejectedByCustomer_NotDone')
        invoice_tasks = InvoiceTask.objects.filter(
            Q(invoice=invoice, status__code__in=allowed_statuses) |
            Q(invoice=invoice, status__parent_code=rejected_notdone_parent)
        ).select_related('task', 'status')

        # Get rejection reasons for sales manager
        rejection_parent = ConstValue.objects.get(code='InvoiceTaskStatus_RejectedByCustomer')
        rejection_reasons = ConstValue.objects.filter(parent_code=rejection_parent)

        # Prepare data for template
        tasks_data = []
        for t in invoice_tasks:
            latest_rejection = RejectionDetails.objects.filter(
                invoice_task=t
            ).order_by('-created_at').first()
            tasks_data.append({
                'task_id': t.task.task_id,
                'task_title': t.task.task_title,
                'task_type': t.task.task_kind.value,
                'project': t.task.project.project_name,
                'team': t.task.project.team_code.team_code,
                'real_work_hours': t.task.real_work_hours_display,
                'invoice_work_hours': t.invoice_work_hours_display,
                'status': t.status.value,
                'status_code': t.status.code,
                'rejection_title': latest_rejection.const_value.value if latest_rejection else '',
                'rejection_explanation': latest_rejection.explanation if latest_rejection else '',
            })

        return render(request, "CorpIssue/sales_manager_form.html", {
            'corp_name': corp.corp_name,
            'corp_code': corp.corp_code,
            'version_number': version_number,
            'invoice': invoice,
            'show_returned_to_pm_table': True,
            'tasks_data': tasks_data,
            'rejection_reasons': rejection_reasons,
            'help_doc': 'CorpIssue/Help Docs/Help 3.pdf'
        })

    invoices = Invoice.objects.filter(id=invoice_id)
    return render(request, "CorpIssue/sales_manager_form.html", {
        'corp_name': corp.corp_name,
        'corp_code': corp.corp_code,
        'version_number': version_number,
        'invoice': invoice,
        'invoices': invoices,
        'help_doc': 'CorpIssue/Help Docs/Help 3.pdf'
    })

@csrf_exempt
@permission_control
def next_stage(request, invoice_id):
    """Handle moving invoice to next stage."""
    if request.method == 'POST':
        try:
            try:
                invoice = Invoice.objects.get(id=invoice_id)
            except Invoice.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'صورت حساب مورد نظر یافت نشد.'})

            token = find_token(request)
            user_nationalcode = get_token_data(token, "user_NationalCode")

            if invoice.status.code == 'InvoiceStatus_SentToSalesManager':
                # Update status to SentToCustomer
                new_status = ConstValue.objects.get(code='InvoiceStatus_SentToCustomer')
                invoice.status = new_status
                invoice.save()
                return JsonResponse({'success': True})
                
            elif invoice.status.code == 'InvoiceStatus_SentToCustomer':
                try:
                    # Get parent for rejection reasons
                    rejection_parent = ConstValue.objects.filter(code='InvoiceTaskStatus_RejectedByCustomer').first()
                    returned_to_pm_status = ConstValue.objects.get(code='InvoiceStatus_ReturnedToProjectManager')
                    rejection_status = rejection_parent
                    parent_id = rejection_parent.id if rejection_parent else None

                    # Process tasks from uploaded Excel
                    invoice_tasks = InvoiceTask.objects.filter(invoice=invoice)
                    excel_file = InvoiceExcel.objects.filter(invoice=invoice).latest('upload_date')
                    team_managers = set()
                    
                    if excel_file and parent_id:
                        df = pd.read_excel(excel_file.file_path)

                        for _, row in df.iterrows():
                            task_id = row['شناسه تسک']
                            rejection_reason = row.get('دلیل عدم تایید')
                            insurance_comments = [row[col] for col in df.columns if col.startswith('نظر شرکت بیمه')]
                            latest_insurance_comment = ''
                            for comment in reversed(insurance_comments):
                                if isinstance(comment, str) and comment.strip():
                                    latest_insurance_comment = comment
                                    break

                            if rejection_reason and not pd.isna(rejection_reason):
                                invoice_task = invoice_tasks.get(task__task_id=task_id)
                                rejection_reason_const = ConstValue.objects.filter(
                                    parent_code=rejection_parent,
                                    value=rejection_reason
                                ).first()
                                if not rejection_reason_const:
                                    rejection_reason_const = rejection_status

                                # Create rejection details
                                RejectionDetails.objects.create(
                                    const_value=rejection_reason_const,
                                    invoice_task=invoice_task,
                                    explanation=latest_insurance_comment if latest_insurance_comment else '',
                                    rejected_by=user_nationalcode,
                                    created_by=user_nationalcode
                                )

                                # Update task status
                                invoice_task.status = rejection_status
                                invoice_task.save()

                                # Add team manager to set
                                team_manager = invoice_task.task.project.team_code.manager
                                if team_manager:
                                    team_managers.add(team_manager)

                    # Update invoice status
                    invoice.status = returned_to_pm_status
                    invoice.save()

                    # Send document to all team managers with rejected tasks
                    if team_managers:
                        send_document(
                            doc_id=invoice.doc_id,
                            sender=user_nationalcode,
                            inbox_owners=list(team_managers)
                        )

                    return JsonResponse({'success': True})

                except ConstValue.DoesNotExist as e:
                    logger.error(f"ConstValue not found: {str(e)}")
                    return JsonResponse({'success': False, 'error': f'Status not found: {str(e)}'})
                except InvoiceExcel.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'No uploaded file found'})
                except Exception as e:
                    logger.error(f"Error processing next stage: {str(e)}")
                    return JsonResponse({'success': False, 'error': str(e)})


                
            return JsonResponse({'success': False, 'error': 'Invalid invoice status'})
            
        except Exception as e:
            logger.error(f"Error in next_stage: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
@permission_control
def approve_task_sales_manager(request, task_id):
    # Set status to ApprovedBySalesManager
    try:
        invoice_task = InvoiceTask.objects.get(task_id=task_id)
        approved_status = ConstValue.objects.get(code='InvoiceTaskStatus_ApprovedBySalesManager')
        invoice_task.status = approved_status
        invoice_task.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@permission_control
def reject_task_sales_manager(request, task_id):
    # Set status to RejectedBySalesManager, create RejectionDetails
    try:
        data = json.loads(request.body)
        reason_id = data.get('reason_id')
        explanation = data.get('explanation')
        invoice_task = InvoiceTask.objects.get(task_id=task_id)
        rejected_status = ConstValue.objects.get(code='InvoiceTaskStatus_RejectedBySalesManager')
        rejection_reason = ConstValue.objects.get(id=reason_id)
        # Create rejection details
        RejectionDetails.objects.create(
            const_value=rejection_reason,
            invoice_task=invoice_task,
            explanation=explanation,
            rejected_by=request.user.username,
            created_by=request.user.username
        )
        invoice_task.status = rejected_status
        invoice_task.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@permission_control
def next_stage_sales_manager(request, invoice_id):
    # Approve all visible tasks, set invoice status to SentToSalesManager
    try:
        invoice = Invoice.objects.get(id=invoice_id)
        approved_status = ConstValue.objects.get(code='InvoiceTaskStatus_ApprovedBySalesManager')
        InvoiceTask.objects.filter(
            invoice=invoice,
            status__code__in=[
                'InvoiceTaskStatus_RejectedByCustomer',
                'InvoiceTaskStatus_ApprovedByProjectManager',
                'InvoiceTaskStatus_RejectedBySalesManager',
                'InvoiceTaskStatus_ApprovedBySalesManager'
            ]
        ).update(status=approved_status)
        sent_status = ConstValue.objects.get(code='InvoiceStatus_SentToSalesManager')
        invoice.status = sent_status
        invoice.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
@csrf_exempt
@permission_control
def approve_all_and_send(request, invoice_id):
    """Approve all tasks and send to customer."""
    if request.method == 'POST':
        try:
            invoice = Invoice.objects.get(id=invoice_id)
            
            # Get the necessary statuses
            approved_status = ConstValue.objects.get(code='InvoiceTaskStatus_ApprovedBySalesManager')
            sent_to_sales_manager_status = ConstValue.objects.get(code='InvoiceStatus_SentToSalesManager')
            
            # Update all eligible tasks
            InvoiceTask.objects.filter(
                invoice=invoice,
                status__code__in=[
                    'InvoiceTaskStatus_ApprovedByProjectManager',
                    'InvoiceTaskStatus_RejectedBySalesManager',
                    'InvoiceTaskStatus_RejectedByCustomer'
                ]
            ).update(status=approved_status)
            
            # Update invoice status
            invoice.status = sent_to_sales_manager_status
            invoice.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})